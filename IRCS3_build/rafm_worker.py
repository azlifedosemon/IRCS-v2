import sys
import os
import pandas as pd
from openpyxl import load_workbook

def read_period0_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}
    data = []
    for r in rows:
        if r[idx.get('period')] == 0:
            data.append((r[idx.get('GOC')], r[idx.get('pol_b')], r[idx.get('cov_units')]))
    df = pd.DataFrame(data, columns=['goc', 'pol_b', 'cov_units'])
    df['pol_b']     = pd.to_numeric(df['pol_b'].astype(str).str.replace(',', '.'), errors='coerce')
    df['cov_units'] = pd.to_numeric(df['cov_units'].astype(str).str.replace(',', '.'), errors='coerce')
    return df

def main(run_name, path):
    try:
        df_idr = read_period0_sheet(path, 'extraction_IDR')
        df_usd = read_period0_sheet(path, 'extraction_USD')
        out = pd.concat([df_idr, df_usd], ignore_index=True)

        pickle_path = f"rafm_{run_name}.pkl"
        out.to_pickle(pickle_path)
        print(f"✅ Saved: {pickle_path}")
    except Exception as e:
        print(f"❌ Error while processing {run_name}: {e}")
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python rafm_worker.py <run_name> <excel_path>")
        sys.exit(1)
    
    run_name = sys.argv[1]
    excel_path = sys.argv[2]

    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)

    main(run_name, excel_path)
