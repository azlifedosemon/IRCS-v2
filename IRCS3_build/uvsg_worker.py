import sys
import os
import pandas as pd
from openpyxl import load_workbook

def read_period0_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"❌ Sheet '{sheet_name}' not found in file: {path}")
    
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}

    required_cols = ['period', 'GOC', 'pol_b', 'rv_av_if']
    for col in required_cols:
        if col not in idx:
            raise ValueError(f"❌ Column '{col}' not found in sheet '{sheet_name}'")

    data = []
    for r in rows:
        if r[idx['period']] == 0:
            data.append((r[idx['GOC']], r[idx['pol_b']], r[idx['rv_av_if']]))
    
    df = pd.DataFrame(data, columns=['goc', 'pol_b', 'rv_av_if'])
    df['pol_b']    = pd.to_numeric(df['pol_b'], errors='coerce')
    df['rv_av_if'] = pd.to_numeric(df['rv_av_if'], errors='coerce')
    return df

def main(run_name, path):
    df_combined = pd.DataFrame()
    try:
        df_idr = read_period0_sheet(path, 'extraction_IDR')
        df_usd = read_period0_sheet(path, 'extraction_USD')
        df_combined = pd.concat([df_idr, df_usd], ignore_index=True)

        output_path = f"uvsg_{run_name}.pkl"
        df_combined.to_pickle(output_path)
        print(f"✅ UVSG saved to: {output_path}")
    except Exception as e:
        print(f"❌ Error processing UVSG '{run_name}': {e}")
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python uvsg_worker.py <run_name> <excel_path>")
        sys.exit(1)

    run_name = sys.argv[1]
    file_path = sys.argv[2]

    if not os.path.exists(file_path):
        print(f"❌ File does not exist: {file_path}")
        sys.exit(1)

    main(run_name, file_path)
