import sys
import pandas as pd
from openpyxl import load_workbook

run_name, path = sys.argv[1], sys.argv[2]

def read_period0_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}
    data = []
    for r in rows:
        if r[idx.get('period')] == 0:
            data.append((r[idx.get('GOC')], r[idx.get('pol_b')], r[idx.get('cov_units')]))
    df = pd.DataFrame(data, columns=['goc', 'pol_b', 'cov_units'])
    df['pol_b'] = pd.to_numeric(df['pol_b'], errors='coerce')
    df['cov_units'] = pd.to_numeric(df['cov_units'], errors='coerce')
    return df

# Baca dari kedua sheet: extraction_IDR dan extraction_USD
df_idr = read_period0_sheet(path, 'extraction_IDR')
df_usd = read_period0_sheet(path, 'extraction_USD')
out = pd.concat([df_idr, df_usd], ignore_index=True)

# Simpan ke file pickle
out.to_pickle(f"rafm_{run_name}.pkl")
