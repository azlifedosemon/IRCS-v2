import pandas as pd
from openpyxl import load_workbook
import sys
import os
from glob import glob

def read_period0_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h: i for i, h in enumerate(hdr)}
    data = []
    for r in rows:
        if r[idx['period']] == 0:
            data.append((r[idx['GOC']], r[idx['pol_b']], r[idx['cov_units']]))
    df = pd.DataFrame(data, columns=['goc', 'pol_b', 'cov_units'])
    df['pol_b'] = pd.to_numeric(df['pol_b'].astype(str).str.replace(',', '.'), errors='coerce')
    df['cov_units'] = pd.to_numeric(df['cov_units'].astype(str).str.replace(',', '.'), errors='coerce')
    return df

def main(path, out_pickle):
    df_idr = read_period0_sheet(path, 'extraction_IDR')
    df_usd = read_period0_sheet(path, 'extraction_USD')
    out = pd.concat([df_idr, df_usd], ignore_index=True)
    out.to_pickle(out_pickle)
    print(f"✅ Saved pickle to {out_pickle}")

if __name__ == '__main__':
    script_dir = os.path.dirname(os.path.abspath(__file__))

    if len(sys.argv) == 3:
        run = sys.argv[1]
        path = sys.argv[2]
        out_pickle = os.path.join(script_dir, f"rafm_{run}.pkl")
        main(path, out_pickle)
    else:
        print("⚠️ No arguments detected. Trying to auto-select Excel file...")

        # Cari file Excel di folder yang sama atau subfolder ./Source/Trad relatif terhadap script
        search_paths = [
            os.path.join(script_dir, "**", "Data_Extraction_run*TRAD_Con.xlsx")
        ]
        matching_files = []
        for pattern in search_paths:
            matching_files.extend(glob(pattern, recursive=True))

        if not matching_files:
            print("❌ No matching Excel files found.")
            sys.exit(1)

        path = sorted(matching_files)[-1]  # Ambil file terakhir
        filename = os.path.basename(path)
        run = filename.split("Data_Extraction_")[1].split("TRAD_Con")[0].lower()
        out_pickle = os.path.join(script_dir, f"rafm_{run}.pkl")

        print(f"📄 Auto-selected file: {filename}")
        main(path, out_pickle)
