import sys
import pandas as pd
from openpyxl import load_workbook

run_name, path = sys.argv[1], sys.argv[2]

def read_period0_sheet(path, sheet_name):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    idx = {h:i for i,h in enumerate(hdr)}
    data = []
    for r in rows:
        if r[idx['period']] == 0:
            data.append((r[idx['GOC']], r[idx['pol_b']], r[idx['RV_AV_IF']]))
    df = pd.DataFrame(data, columns=['goc','pol_b','RV_AV_IF'])
    df['pol_b']     = pd.to_numeric(df['pol_b'].astype(str).str.replace(',','.'), errors='coerce')
    df['RV_AV_IF'] = pd.to_numeric(df['RV_AV_IF'].astype(str).str.replace(',','.'), errors='coerce')
    
    return df

# parse both sheets and save
df_idr = read_period0_sheet(path, 'extraction_IDR')
df_usd = read_period0_sheet(path, 'extraction_USD')
out = pd.concat([df_idr, df_usd], ignore_index=True)

# dump to a temp file
out.to_pickle(f"rafm_{run_name}.pkl")