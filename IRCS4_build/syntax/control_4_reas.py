import pandas as pd
import glob
import os
from concurrent.futures import ProcessPoolExecutor
import re
from openpyxl import load_workbook

columns_to_sum_argo = [
    'prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost','cov_units','DAC_COV_UNITS','dac','exp_acq',
    'lrc_cl_ins_dth','lrc_cl_inv_dth','lrc_cl_inv_surr','lrc_cl_inv_mat','lrc_cl_inv_ann'
    ]
columns_to_sum_rafm = ['prm_inc','lrc_cl_ins','cov_units','pv_reins_clm','lrc_cl_ins_dth']
cols_to_compare = ['prm_inc','lrc_cl_ins','cov_units','dac_cov_units','lrc_cl_ins_dth']
target_sheets = ['extraction IDR', 'extraction USD']
global_filter_rafm = None
def process_argo_file(file_path):
    try:
        file_name_argo = os.path.splitext(os.path.basename(file_path))[0]
        wb = load_workbook(file_path, read_only=True, data_only=True)

        if 'Sheet1' not in wb.sheetnames:
            wb.close()
            return None

        sheet = wb['Sheet1']
        rows = sheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            return None

        col_index = {col: i for i, col in enumerate(header) if col in columns_to_sum_argo}
        if not col_index:
            wb.close()
            return None

        sums = {col: 0 for col in col_index}

        for row in rows:
            for col, idx in col_index.items():
                if idx < len(row):
                    val = row[idx]
                    if isinstance(val, (int, float)):
                        sums[col] += val

        wb.close()
        sums['File_Name'] = file_name_argo
        return sums

    except Exception as e:
        print(f"❌ Error processing {file_path}: {e}")
        return None

def process_rafm_file(entry):
    file_path, file_name = entry
    total_sums = {col: 0 for col in columns_to_sum_rafm}

    for sheet_name in target_sheets:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                continue

            sheet = wb[sheet_name]
            rows = sheet.iter_rows(values_only=True)

            max_skip = 20
            header = None
            for _ in range(max_skip):
                raw = next(rows, [])
                cleaned = [str(h).strip().lower() if h else '' for h in raw]
                if 'goc' in cleaned:
                    header = cleaned
                    break

            if not header:
                print(f"⚠️ Kolom 'GOC' tidak ditemukan dalam 20 baris pertama di sheet {sheet_name} file {file_name}, dilewati.")
                wb.close()
                continue

            for _ in range(3):
                peek = next(rows, [])
                if any(peek):
                    data_start = [peek]
                    break
            else:
                data_start = []

            col_index = {}
            for i, col in enumerate(header):
                if col in [c.lower() for c in columns_to_sum_rafm] or col == 'goc':
                    col_index[col] = i

            for row in data_start + list(rows):
                for col in columns_to_sum_rafm:
                    idx = col_index.get(col.lower())
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if isinstance(val, (int, float)) and val != 0:
                            total_sums[col] += val

            wb.close()

        except Exception as e:
            print(f"   ❌ Error processing sheet {sheet_name} file {file_name}: {e}")
            continue

    total_sums['File_Name'] = file_name
    return total_sums

def main(params):
    global columns_to_sum_argo, columns_to_sum_rafm, cols_to_compare, target_sheets

    input_excel = params['input excel']

    code = pd.read_excel(input_excel, sheet_name='Code')
    sign_logic = pd.read_excel(input_excel, sheet_name='Sign Logic')
    control = pd.read_excel(input_excel,sheet_name = 'Control')
    file_path_df = pd.read_excel(input_excel, sheet_name='File Path')
    path_map = dict(zip(file_path_df['Name'].str.lower(), file_path_df['File Path']))

    folder_path_argo = path_map.get('argo', '')
    folder_path_rafm = path_map.get('rafm', '')
    rafm_manual_path = path_map.get('rafm manual', '')

    file_paths_argo = [f for f in glob.glob(os.path.join(folder_path_argo, '*.xlsx')) if not os.path.basename(f).startswith('~$')]
    with ProcessPoolExecutor() as executor:
        summary_rows_argo = list(filter(None, executor.map(process_argo_file, file_paths_argo)))

    cf_argo = pd.DataFrame(summary_rows_argo)
    cf_argo = cf_argo[['File_Name'] + [col for col in cf_argo.columns if col != 'File_Name']]
    cf_argo = cf_argo.rename(columns={'File_Name': 'ARGO File Name', 'DAC_COV_UNITS': 'dac_cov_units'})

    file_paths_rafm = [f for f in glob.glob(os.path.join(folder_path_rafm, '*.xlsx')) if not os.path.basename(f).startswith('~$')]
    file_entries = [(f, os.path.splitext(os.path.basename(f))[0]) for f in file_paths_rafm]

    with ProcessPoolExecutor() as executor:
        results = list(executor.map(process_rafm_file, file_entries))

    summary_rows_rafm = []
    for result in results:
        if result:
            summary_rows_rafm.append(result)

    all_runs = ['11', '21', '31', '41']

    original_data_by_file = {row['File_Name']: row.copy() for row in summary_rows_rafm}

    from collections import defaultdict
    grouped_files = defaultdict(dict)

    for row in summary_rows_rafm:
        file_name = row['File_Name']
        for run in all_runs:
            if f"run{run}" in file_name:
                prefix = file_name.split(f"run{run}")[0]
                grouped_files[prefix][run] = file_name
                break 

    for prefix, run_map in grouped_files.items():
        present_runs = sorted(run_map.keys()) 

        for target_run in present_runs:
            target_file = run_map[target_run]


            runs_to_sum = [r for r in all_runs if r <= target_run and r in run_map]

            total_sum = {col: 0 for col in columns_to_sum_rafm }
            for run in runs_to_sum:
                file = run_map[run]
                data_row = original_data_by_file[file]
                for col in total_sum:
                    total_sum[col] += data_row.get(col, 0)

            for i, row in enumerate(summary_rows_rafm):
                if row['File_Name'] == target_file:
                    for col in total_sum:
                        summary_rows_rafm[i][col] = total_sum[col]
                    break

    cf_rafm_1 = pd.DataFrame(summary_rows_rafm)
    cf_rafm_1 = cf_rafm_1[['File_Name'] + [col for col in cf_rafm_1.columns if col != 'File_Name']]
    cf_rafm_1 = cf_rafm_1.rename(columns={'File_Name': 'RAFM File Name'})
    cf_rafm_1 = cf_rafm_1.groupby('RAFM File Name', as_index=False).first()
    cf_rafm_merge = pd.merge(code, cf_rafm_1, on="RAFM File Name", how="left").fillna(0)

    numeric_cols = cf_rafm_merge.select_dtypes(include='number').columns
    sum_rows = cf_rafm_merge[cf_rafm_merge['RAFM File Name'].str.contains("SUM_", na=False)]

    for idx, row in sum_rows.iterrows():
        rafm_value = row['RAFM File Name']

        if 'SUM_' in rafm_value:
            keyword = rafm_value.split('SUM_')[-1]

            pattern = re.escape(keyword).replace("-", "[-_]?")
            matched_rows = cf_rafm_merge[cf_rafm_merge['ARGO File Name'].fillna('').str.contains(pattern, case=False, regex=True)]

            total_values = matched_rows[numeric_cols].sum()

            for col in numeric_cols:
                cf_rafm_merge.at[idx, col] = total_values[col]

    cf_rafm = cf_rafm_merge.drop(columns=['ARGO File Name'])
    cf_rafm['dac_cov_units'] = cf_rafm['cov_units']

    rafm_manual = pd.read_excel(rafm_manual_path, sheet_name = 'Sheet1',engine = 'openpyxl')
    rafm_manual = rafm_manual.drop (columns = ['No'])
    rafm_manual = rafm_manual.fillna(0)

    cf_rafm_2 = cf_rafm.groupby('RAFM File Name', as_index=False).first()
    rafm_manual_1 = rafm_manual.groupby('RAFM File Name', as_index=False).first()
    rafm = pd.merge(cf_rafm_2, rafm_manual_1, on='RAFM File Name', how="left", suffixes=('_cf', '_manual')).fillna(0)
    for col in cols_to_compare:
        col_cf = f"{col}_cf"             
        col_manual = f"{col}_manual" 

        if col_cf in rafm.columns and col_manual in rafm.columns:
            rafm[f'{col}'] = rafm[col_cf] - rafm[col_manual] 
    cols_final = ['RAFM File Name']+ [f'{col}' for col in cols_to_compare]
    rafm = rafm[cols_final]
    rafm = rafm.fillna(0)
    rafm_merged = pd.merge(code, rafm.groupby('RAFM File Name', as_index=False).first(), on="RAFM File Name", how="left").fillna(0)
    final = pd.merge(rafm_merged, cf_argo, on='ARGO File Name', how="left", suffixes=('_merged', '_argo'))

    for col in cols_to_compare:
        col_argo = f"{col}_argo"
        col_merged = f"{col}_merged"
        if col_argo in final.columns and col_merged in final.columns:
            final[f'{col}_diff'] = final[col_merged] - final[col_argo]
        else:
            print(f"Skipped {col} because columns not found: {col_argo}, {col_merged}")

    cols_final = ['RAFM File Name', 'ARGO File Name'] + [f'{col}_diff' for col in cols_to_compare]
    final = final[cols_final].fillna(0)
    logic_row = sign_logic.iloc[0]

    valid_cols = [col for col in logic_row.index if col in cf_argo.columns]
    def check_sign(val, logic_sign):
        if pd.isna(val):
            return 0
        if logic_sign == 1:
            return 1 if val < 0 else 0
        elif logic_sign == "-":
            return 0  
        elif logic_sign == -1:
            return 1 if val > 0 else 0 
        return 0 

    check_sign_summary_row = {
        col: cf_argo[col].apply(lambda val: check_sign(val, logic_row[col])).sum()
        for col in valid_cols
    }

    for col in cf_argo.columns:
        if col not in check_sign_summary_row:
            check_sign_summary_row[col] = None
    check_sign_summary = pd.DataFrame([check_sign_summary_row])[cf_argo.columns]
    cf_argo = pd.concat([cf_argo, check_sign_summary], ignore_index=True)
    check_sign_total = sum(val for val in check_sign_summary_row.values() if isinstance(val, (int, float)))
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = check_sign_total
    index_labels = list(range(1, len(cf_argo))) + ['check sign']
    cf_argo.insert(0, 'No', index_labels)
    cf_argo = pd.concat([cf_argo, sign_logic], ignore_index=True)
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = 'Sign Logic'
    index_labels_rafm = list(range(1, len(cf_rafm) + 1))
    cf_rafm.insert(0, 'No', index_labels_rafm)
    index_labels_manual= list(range(1, len(rafm_manual)+1))
    rafm_manual.insert(0, 'No', index_labels_manual)
    index_labels_final= list(range(1, len(final)+1))
    final.insert(0, 'No', index_labels_final)

    control['check sign'] = ''
    control['result'] = ''

    val_year_idx = control[control.iloc[:, 0] == 'Val Year'].index
    if not val_year_idx.empty:
        idx = val_year_idx[0]
        control.at[idx, 'check sign'] = 'Check Sign'
        control.at[idx, 'result'] = check_sign_total

    return {
        'Control':control,
        'Code':code,
        "CF ARGO REAS": cf_argo,
        "RAFM Output REAS": cf_rafm,
        "RAFM Manual REAS" : rafm_manual,
        "Checking Summary REAS": final
    }
if __name__ == '__main__':
    import multiprocessing
    multiprocessing.freeze_support()

