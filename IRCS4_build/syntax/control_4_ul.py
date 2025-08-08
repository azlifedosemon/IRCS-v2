import pandas as pd
import glob
import os
from concurrent.futures import ProcessPoolExecutor
import re
from openpyxl import load_workbook


columns_to_sum_argo = [
    'prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost',
    'cov_units','dac_cov_units','dac','nattr_exp_acq','nattr_exp_inv',
    'nattr_exp_maint','nattr_exp','tab_dedn','u_sar','pv_r_exp_m','pv_surr','pv_pw_n','pv_clm_surr_pw_n',
    'lrc_cl_ins_dth','lrc_cl_inv_dth','lrc_cl_inv_surr','lrc_cl_inv_mat',
    'clm_base','clm_pro','clm_hth','nattr_exp_maint_inv'
]

columns_to_sum_rafm = [
    'period','prm_inc','lrc_cl_ins','lrc_cl_inv','r_exp_m','r_acq_cost',
    'dac','nattr_exp_acq','nattr_exp_inv','nattr_exp_maint',
    'tab_dedn','lrc_cl_ins_dth','lrc_cl_inv_dth',
    'lrc_cl_inv_surr','lrc_cl_inv_mat','clm_base','clm_pro','clm_hth','tab_ph'
]

additional_columns = ['pv_pw_n','cov_units', 'u_sar', 'pv_r_exp_m', 'pv_surr']
target_sheets = ['extraction_IDR', 'extraction_USD']
global_filter_rafm = None
all_runs = ['11', '21', '31', '41']


def process_argo_file(file_path):
    file_name_argo = os.path.splitext(os.path.basename(file_path))[0]
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb['Sheet1']
    rows = sheet.iter_rows(values_only=True)
    header = next(rows)
    col_index = {col: i for i, col in enumerate(header) if col in columns_to_sum_argo}
    sums = {col: 0 for col in col_index}
    for row in rows:
        for col, idx in col_index.items():
            if idx < len(row):
                val = row[idx]
                if isinstance(val, (int, float)):
                    sums[col] += val
    wb.close()
    sums['File_Name'] = file_name_argo
    return sums


def process_rafm_file(args):
    file_path, file_name, filter_df = args
    match = filter_df[filter_df['File Name'] == file_name]
    if match.empty:
        return None

    total_sums = {col: 0 for col in columns_to_sum_rafm}
    additional_sums = {col: 0 for col in additional_columns}

    speed = int(match['Speed Duration'].values[0])
    exclude = str(match['Exclude Year'].values[0])
    include = str(match['Include Year'].values[0])

    for sheet_name in target_sheets:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                continue
            actual_sheetnames = [s.strip().lower() for s in wb.sheetnames]
            target_lower = sheet_name.lower()

            if target_lower not in actual_sheetnames:
                wb.close()
                continue
            matched_sheet = wb.sheetnames[actual_sheetnames.index(target_lower)]
            sheet = wb[matched_sheet]
            rows = sheet.iter_rows(values_only=True)
            header = next(rows)

            col_index = {}
            for i, col in enumerate(header):
                col_name = str(col).strip().lower() if col else ''
                if col_name in [c.lower() for c in columns_to_sum_rafm + additional_columns] or col_name == 'goc':
                    col_index[col_name] = i

            if 'goc' not in col_index:
                continue

            for row in rows:
                val_goc = str(row[col_index['goc']]) if col_index['goc'] < len(row) else ''
                period_idx = col_index.get('period')
                period_value = row[period_idx] if period_idx is not None and period_idx < len(row) else None

                skip_row = False
                if include != '-' and exclude != '-':
                    if include not in val_goc or exclude in val_goc:
                        skip_row = True
                elif include != '-':
                    if include not in val_goc:
                        skip_row = True
                elif exclude != '-':
                    if exclude in val_goc:
                        skip_row = True

                if skip_row:
                    continue

                if period_value is not None and period_value > speed:
                    for col in columns_to_sum_rafm:
                        idx = col_index.get(col.lower())
                        if idx is not None and idx < len(row):
                            val = row[idx]
                            if isinstance(val, (int, float)) and val != 0:
                                total_sums[col] += val

                if period_value is not None and period_value >= 0:
                    for col in additional_columns:
                        idx = col_index.get(col.lower())
                        if idx is not None and idx < len(row):
                            val = row[idx]
                            if isinstance(val, (int, float)) and val != 0:
                                additional_sums[col] += val
            wb.close()

        except:
            continue

    total_sums['File_Name'] = file_name
    additional_sums['File_Name'] = file_name
    return total_sums, additional_sums


def main(params):
    global global_filter_rafm

    input_excel = params['input excel']

    code = pd.read_excel(input_excel, sheet_name='Code')
    sign_logic = pd.read_excel(input_excel, sheet_name='Sign Logic')
    control = pd.read_excel(input_excel, sheet_name='Control')
    file_path_df = pd.read_excel(input_excel, sheet_name='File Path')
    global_filter_rafm = pd.read_excel(input_excel, sheet_name='Filter RAFM')

    path_map = dict(zip(file_path_df['Name'].str.lower(), file_path_df['File Path']))
    folder_path_argo = path_map.get('argo', '')
    folder_path_rafm = path_map.get('rafm', '')
    rafm_manual_path = path_map.get('rafm manual', '')

    # === ARGO ===
    file_paths_argo = [
        f for f in glob.glob(os.path.join(folder_path_argo, '*.xlsx'))
        if not os.path.basename(f).startswith('~$')
    ]

    with ProcessPoolExecutor() as executor:
        summary_rows_argo = list(executor.map(process_argo_file, file_paths_argo))

    cf_argo = pd.DataFrame(summary_rows_argo)
    cf_argo = cf_argo.rename(columns={'File_Name': 'ARGO File Name'})
    cf_argo = pd.merge(code,cf_argo, on = 'ARGO File Name', how = 'left')
    
    columns_to_drop = []
    if 'RAFM File Name' in cf_argo.columns:
        columns_to_drop.append('RAFM File Name')
    if 'UVSG File Name' in cf_argo.columns:
        columns_to_drop.append('UVSG File Name')
    if columns_to_drop:
        cf_argo = cf_argo.drop(columns=columns_to_drop)
    
    if 'ARGO File Name' in cf_argo.columns:
        cols = ['ARGO File Name'] + [col for col in cf_argo.columns if col != 'ARGO File Name']
        cf_argo = cf_argo[cols]

    # === RAFM ===
    file_paths_rafm = [f for f in glob.glob(os.path.join(folder_path_rafm, '*.xlsx')) if not os.path.basename(f).startswith('~$')]
    file_entries = [
    (f, os.path.splitext(os.path.basename(f))[0], global_filter_rafm)
    for f in file_paths_rafm
    ]

    with ProcessPoolExecutor() as executor:
        results = list(executor.map(process_rafm_file, file_entries))

    summary_rows_rafm = []
    additional_summary_rows = []
    for result in results:
        if result:
            total_sums, additional_sums = result
            summary_rows_rafm.append(total_sums)
            additional_summary_rows.append(additional_sums)

    combined_summary = []
    for main_row, add_row in zip(summary_rows_rafm, additional_summary_rows):
        combined_row = {**main_row, **add_row}
        combined_summary.append(combined_row)

    all_runs = ['11', '21', '31', '41']

    original_data_by_file = {row['File_Name']: row.copy() for row in combined_summary}

    from collections import defaultdict
    grouped_files = defaultdict(dict)

    for row in combined_summary:
        file_name = row['File_Name']
        for run in all_runs:
            if f"run{run}" in file_name:
                prefix = file_name.split(f"run{run}")[0]
                grouped_files[prefix][run] = file_name
                break 

    for prefix, run_map in grouped_files.items():
        present_runs = sorted(run_map.keys()) 

        for target_run in present_runs:
            target_file = run_map[target_run]


            runs_to_sum = [r for r in all_runs if r <= target_run and r in run_map]

            total_sum = {col: 0 for col in columns_to_sum_rafm + additional_columns}
            for run in runs_to_sum:
                file = run_map[run]
                data_row = original_data_by_file[file]
                for col in total_sum:
                    total_sum[col] += data_row.get(col, 0)

            for i, row in enumerate(combined_summary):
                if row['File_Name'] == target_file:
                    for col in total_sum:
                        combined_summary[i][col] = total_sum[col]
                    break

    cf_rafm = pd.DataFrame(combined_summary).rename(columns={'File_Name': 'RAFM File Name'})
    cf_rafm_merge = pd.merge(code, cf_rafm, on="RAFM File Name", how="left").fillna(0)

    sum_rows = cf_rafm_merge[cf_rafm_merge['RAFM File Name'].str.contains("SUM_", na=False)]
    numeric_cols = cf_rafm_merge.select_dtypes(include='number').columns

    for idx, row in sum_rows.iterrows():
        keyword = row['RAFM File Name'].split('SUM_')[-1]
        pattern = re.escape(keyword).replace("-", "[-_]?")
        matched = cf_rafm_merge[cf_rafm_merge['ARGO File Name'].fillna('').str.contains(pattern, case=False, regex=True)]
        totals = matched[numeric_cols].sum()
        for col in numeric_cols:
            cf_rafm_merge.at[idx, col] = totals[col]

    columns_to_drop = []
    if 'ARGO File Name' in cf_rafm_merge.columns:
        columns_to_drop.append('ARGO File Name')
    if 'period' in cf_rafm_merge.columns:
        columns_to_drop.append('period')
    if 'UVSG File Name' in cf_rafm_merge.columns:
        columns_to_drop.append('UVSG File Name')
    if columns_to_drop:
        cf_rafm = cf_rafm_merge.drop(columns=columns_to_drop)
    else:
        cf_rafm = cf_rafm_merge.copy()
    
    cf_rafm['dac'] = -cf_rafm['r_acq_cost']
    cf_rafm['nattr_exp'] = cf_rafm[['nattr_exp_acq', 'nattr_exp_inv', 'nattr_exp_maint']].sum(axis=1)
    cf_rafm['pv_clm_surr_pw_n'] = cf_rafm[['pv_surr', 'pv_pw_n']].sum(axis=1)
    cf_rafm['nattr_exp_maint_inv'] = cf_rafm[['nattr_exp_inv', 'nattr_exp_maint']].sum(axis=1)
    cf_rafm['dac_cov_units'] = cf_rafm['cov_units']

    rafm_manual = pd.read_excel(rafm_manual_path, sheet_name='Sheet1').drop(columns=['No','Update (Y/N)','Shift Dur','Cohort'])
    rafm_manual = rafm_manual.rename(columns={'c_sar': 'u_sar'}).fillna(0)
    final = code.copy()
    for col in columns_to_sum_argo:
        if col not in code.columns:
            final[col] = pd.NA
    logic_row = sign_logic.iloc[0]

    mapping_code = global_filter_rafm.drop(columns = {'File Name'})
    mapping = pd.concat([code,mapping_code], axis = 1)
    global_filter_rafm = global_filter_rafm.groupby('File Name', as_index = False).first()
    global_filter_rafm = global_filter_rafm.rename(columns = {'File Name':'RAFM File Name'})
    cf_rafm = pd.merge(cf_rafm,global_filter_rafm,on = 'RAFM File Name', how = 'left')
    logic_row = sign_logic.iloc[0]

    valid_cols = [col for col in logic_row.index if col in cf_argo.columns]
    def check_sign(val, logic_sign):
        if pd.isna(val):
            return 0
        if logic_sign == 1:
            return 1 if val < 0 else 0
        elif logic_sign == "-":
            return 0
        elif logic_sign == -1:
            return 1 if val > 0 else 0 
        return 0 

    check_sign_summary_row = {
        col: cf_argo[col].apply(lambda val: check_sign(val, logic_row[col])).sum()
        for col in valid_cols
    }

    for col in cf_argo.columns:
        if col not in check_sign_summary_row:
            check_sign_summary_row[col] = None
    check_sign_summary = pd.DataFrame([check_sign_summary_row])[cf_argo.columns]
    cf_argo = pd.concat([cf_argo, check_sign_summary], ignore_index=True)
    check_sign_total = sum(val for val in check_sign_summary_row.values() if isinstance(val, (int, float)))
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = check_sign_total
    index_labels = list(range(1, len(cf_argo))) + ['check sign']
    cf_argo.insert(0, 'No', index_labels)
    cf_argo = pd.concat([cf_argo, sign_logic], ignore_index=True)
    cf_argo.loc[cf_argo.index[-1], 'ARGO File Name'] = 'Sign Logic'
    index_labels_manual= list(range(1, len(rafm_manual)+1))
    rafm_manual.insert(0, 'No', index_labels_manual)
    index_labels_final= list(range(1, len(final)+1))
    final.insert(0, 'No', index_labels_final)

    control['check sign'] = ''
    control['result'] = ''

    val_year_idx = control[control.iloc[:, 0] == 'Val Year'].index
    if not val_year_idx.empty:
        idx = val_year_idx[0]
        control.at[idx, 'check sign'] = 'Check Sign'
        control.at[idx, 'result'] = check_sign_total
    if 'RAFM File Name' in cf_rafm.columns:
        last_3_cols = cf_rafm.columns[-3:].tolist()
        other_cols = [col for col in cf_rafm.columns if col not in last_3_cols and col != 'RAFM File Name']
        cf_rafm = cf_rafm[['RAFM File Name'] + last_3_cols + other_cols]
    index_labels_rafm = list(range(1, len(cf_rafm)+1))
    cf_rafm.insert(0, 'No', index_labels_rafm)
    return {
        'Control':control,
        'Code':mapping,
        "CF ARGO AZUL": cf_argo,
        "RAFM Output AZUL": cf_rafm,
        "RAFM Output Manual": rafm_manual,
        "Checking Summary AZUL": final
    }



if __name__ == '__main__':
    import multiprocessing
    multiprocessing.freeze_support()